"""Gerador de relatórios Excel para notas de provas."""
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.models.entities.exams import Exams
from src.models.entities.student_answers import StudentAnswer
from src.models.entities.exam_questions import ExamQuestion
from src.models.entities.user import User

from src.core.logging_config import get_logger

logger = get_logger(__name__)


def generate_grades_report(
    db: Session,
    exam: Exams,
    student_answers: List[StudentAnswer]
) -> str:
    """
    Gera um relatório Excel com as notas dos alunos.
    
    Args:
        db: Sessão do banco de dados
        exam: Objeto da prova
        student_answers: Lista de respostas dos alunos
        
    Returns:
        str: Caminho relativo do arquivo Excel gerado
    """
    
    # Criar diretório de relatórios se não existir
    reports_dir = Path("data/reports")
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Nome do arquivo com timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"notas_{exam.uuid}_{timestamp}.xlsx"
    filepath = reports_dir / filename
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas"
    
    # Buscar questões da prova (ordenadas)
    questions = db.query(ExamQuestion).filter(
        ExamQuestion.exam_uuid == exam.uuid
    ).order_by(ExamQuestion.question_order).all()
    
    # Estilo para cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = ["Aluno"]
    for i, question in enumerate(questions, 1):
        max_score = question.max_score or 10.0
        headers.append(f"Q{i} (/{max_score:.1f})")
    headers.append("Total")
    
    # Escrever cabeçalhos
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Agrupar respostas por aluno
    student_grades = {}
    for answer in student_answers:
        student_uuid = answer.student_uuid
        if student_uuid not in student_grades:
            # Buscar nome do aluno
            student = db.query(User).filter(User.uuid == student_uuid).first()
            student_name = student.name if student else str(student_uuid)
            student_grades[student_uuid] = {
                "name": student_name,
                "questions": {},
                "total": 0.0
            }
        
        # Encontrar índice da questão
        question_idx = next(
            (i for i, q in enumerate(questions) if q.uuid == answer.question_uuid),
            None
        )
        
        if question_idx is not None:
            # Usar nota final se existir, senão nota da IA
            score = answer.final_score if answer.final_score is not None else answer.ai_score
            student_grades[student_uuid]["questions"][question_idx] = score or 0.0
            student_grades[student_uuid]["total"] += score or 0.0
    
    # Escrever dados dos alunos
    row_idx = 2
    for student_data in sorted(student_grades.values(), key=lambda x: x["name"]):
        # Nome do aluno
        ws.cell(row=row_idx, column=1, value=student_data["name"])
        
        # Notas por questão
        for q_idx in range(len(questions)):
            score = student_data["questions"].get(q_idx, 0.0)
            cell = ws.cell(row=row_idx, column=q_idx + 2, value=score)
            cell.alignment = Alignment(horizontal="center")
        
        # Total
        total_cell = ws.cell(row=row_idx, column=len(questions) + 2, value=student_data["total"])
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="center")
        
        row_idx += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30  # Nome do aluno
    for col_idx in range(2, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 12
    
    # Salvar arquivo
    wb.save(filepath)
    
    logger.info("Relatório Excel gerado: %s", filepath)
    
    # Retornar caminho relativo para download
    return str(filepath.relative_to(Path.cwd()))
